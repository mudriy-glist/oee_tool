# Author: Peter Kulisov
# Copyright: Peter Kulisov <peter.kulisov@gmail.com>
# If there are any issues contact me on the email above.
#
# Version 1.0
# Date: 11/08/2022
from openpyxl import Workbook
import datetime

def report_generate(data_dict, unique_machines, target_time):

    #instance for Workbook class
    wb = Workbook()

    #instance for sheet
    ws = wb.create_sheet("DATA")

    #Headers
    ws['A1'].value = "ID"
    ws['B1'].value = "Machine"
    ws['C1'].value = "Date"
    ws['D1'].value = "Qty"
    ws['E1'].value = "Total Time"
    ws['F1'].value = "Time lost"
    ws['G1'].value = "OEE time"

    i = 0
    #Data from data_dict iterates and inputs
    for u in unique_machines:
        for k, v in data_dict[u].items():
            ws[f'A{i + 2}'].value = i + 1 #ID
            ws[f'B{i + 2}'].value = u #Machine
            ws[f'C{i + 2}'].value = k #date
            ws[f'D{i + 2}'].value = v[0] #qty
            ws[f'E{i + 2}'].value = v[1] #total_time
            ws[f'F{i + 2}'].value = f"{round((v[1] - target_time)*-1, 2)}" # Time loses
            ws[f'G{i + 2}'].value = f"{round((v[1] / target_time) * 100, 2)}%" # OEE time
            i += 1
    
    #remove firts empty sheet
    wb.remove(wb['Sheet'])

    #get todays date
    now = datetime.datetime.now()
    now_date_only = now.date()
    format = "%d-%m-%Y"
    filename_date = now_date_only.strftime(format)

    #save file
    wb.save(filename=f"OEE_Report{filename_date}.xlsx")