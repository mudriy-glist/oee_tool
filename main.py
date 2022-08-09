from openpyxl import Workbook
import pandas as pd
import datetime

df = pd.read_excel("./export_oee.xlsx")
df = pd.DataFrame({
    "machine" : df["TSWCID"],
    "part_no" : df["WOPartNo"],
    "wo" : df["TSWONo"],
    "op_no" : df["TSOperNo"],
    "total_time_hour" : df["TSTotTimeDec"],
    "qty" : df["TSQuantity"],
    "start_time" : df["TSStartTime"],
    "finish_time" : df["TSFinishTime"],
    "employee_id" : df["TSEmployeeID"],
    "op_complete" : df["TSOpComplete"],
    "date" : None
})
df["date"] = [d.date() for d in df["start_time"]]
data_dict = {}
unique_machines = list(df['machine'].unique())
for u in unique_machines:
    data_dict[u] = {}
    for i in range(len(df)):
        if df.loc[i, "machine"] == u:
            date = df.loc[i, "date"]
            add_list = [df.loc[i, "qty"], df.loc[i, "total_time_hour"]]
            if f"{date}" not in data_dict[u]:
                data_dict[u][f"{date}"] = add_list
            else:
                exch_list = data_dict[u][f"{date}"]
                data_dict[u][f"{date}"] = [add_list[0] + exch_list[0], add_list[1] + exch_list[1]]
                
wb = Workbook()
ws = wb.create_sheet("DATA")
i = 0
ws['A1'].value = "ID"
ws['B1'].value = "Machine"
ws['C1'].value = "Date"
ws['D1'].value = "Qty"
ws['E1'].value = "Total Time"
ws['F1'].value = "Time lost"
ws['G1'].value = "OEE time"

target_time = 18.316

for u in unique_machines:
    for k, v in data_dict[u].items():
        ws[f'A{i + 2}'].value = i + 1 #ID
        ws[f'B{i + 2}'].value = u #Machine
        ws[f'C{i + 2}'].value = k #date
        ws[f'D{i + 2}'].value = v[0] #qty
        ws[f'E{i + 2}'].value = v[1] #total_time
        ws[f'F{i + 2}'].value = f"{round((v[1] - target_time)*-1, 2)}" # Time loses
        ws[f'G{i + 2}'].value = f"{round((v[1] / target_time) * 100, 2)}%" # OEE time
        i += 1

length_f = len(ws['A'])
wb.remove(wb['Sheet'])
now = datetime.datetime.now()
now_date_only = now.date()
format = "%d-%m-%Y"
filename_date = now_date_only.strftime(format)
wb.save(filename=f"OEE_Report{filename_date}.xlsx")